"""
Report generation functionality for Semgrep findings.
"""

import csv
import datetime
import logging
import json
from pathlib import Path
from typing import List, Dict, Any

import pandas as pd
import plotly.express as px
from fpdf import FPDF
from jinja2 import Environment, PackageLoader, select_autoescape
from openpyxl.utils import get_column_letter

from .api import Finding
from .config import ReportConfig

logger = logging.getLogger("semgrep_reporter")


class ReportGenerator:
    """Generates reports in various formats from Semgrep findings."""

    def __init__(self, config: ReportConfig, output_dir: Path):
        self.config = config
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize Jinja2 environment
        self.jinja_env = Environment(
            loader=PackageLoader("semgrep_reporter", "templates"),
            autoescape=select_autoescape(['html', 'xml'])
        )
        logger.debug(f"ReportGenerator initialized with output directory: {output_dir}")

    def generate_reports(self, findings: List[Finding]) -> None:
        """
        Generate reports in all configured formats.

        Args:
            findings: List of Semgrep findings to include in the reports
        """
        logger.info(f"Generating reports for {len(findings)} findings in formats: {self.config.output_formats}")
        
        # First save the raw findings data
        try:
            raw_findings = [finding.raw_response for finding in findings if hasattr(finding, 'raw_response')]
            if raw_findings:
                raw_output_path = self.output_dir / "semgrep_raw_findings.json"
                with open(raw_output_path, 'w') as f:
                    json.dump(raw_findings, f, indent=2)
                logger.info(f"Raw API response saved to {raw_output_path}")
        except Exception as e:
            logger.error(f"Error saving raw API response: {e}")
            logger.exception("Raw data save error details")

        # Generate configured report formats
        for format_type in self.config.output_formats:
            try:
                if format_type == "pdf":
                    self._generate_pdf_report(findings)
                elif format_type == "csv":
                    self._generate_csv_report(findings)
                elif format_type == "xlsx":
                    self._generate_excel_report(findings)
                logger.debug(f"Successfully generated {format_type} report")
            except Exception as e:
                logger.error(f"Error generating {format_type} report: {e}")
                logger.exception("Report generation error details")

    def _finding_to_dict(self, finding: Finding) -> Dict[str, Any]:
        """Convert a Finding object to a dictionary with consistent field access."""
        finding_dict = {
            # Core fields
            "check_id": finding.check_id,
            "path": finding.get_path(),
            "line": finding.get_line(),
            "message": finding.get_message(),
            "severity": finding.get_severity(),
            "repository": finding.get_repository_name(),
            
            # Status and triage fields
            "state": getattr(finding, 'state', None),
            "status": getattr(finding, 'status', None),
            "triage_state": getattr(finding, 'triage_state', None),
            "triaged_at": getattr(finding, 'triaged_at', None),
            "triage_comment": getattr(finding, 'triage_comment', None),
            "triage_reason": getattr(finding, 'triage_reason', None),
            "state_updated_at": getattr(finding, 'state_updated_at', None),
            
            # Optional fields
            "commit": finding.commit,
            "scan_date": finding.scan_date,
            "line_of_code_url": finding.line_of_code_url,
        }

        # Add Semgrep UI link if raw_response is available
        if hasattr(finding, 'raw_response') and isinstance(finding.raw_response, dict):
            finding_id = finding.raw_response.get('id')
            finding_dict.update({
                "finding_id": finding_id,
                "semgrep_ui_url": f"https://semgrep.dev/orgs/{self.config.deployment_slug}/findings/{finding_id}" if finding_id else None
            })
        else:
            finding_dict.update({
                "finding_id": None,
                "semgrep_ui_url": None
            })
            
        # Add SCA-specific fields
        finding_dict.update({
            "is_dependency": getattr(finding, 'is_dependency', False),
            "dependency_name": getattr(finding, 'dependency_name', None),
            "dependency_version": getattr(finding, 'dependency_version', None),
            "fixed_version": getattr(finding, 'fixed_version', None),
            "ecosystem": getattr(finding, 'ecosystem', None),
            "cve_ids": getattr(finding, 'cve_ids', []),
            "references": getattr(finding, 'references', []),
            "reachable": getattr(finding, 'reachable', None),
            "reachability_details": getattr(finding, 'reachability_details', None),
            
            # Rule fields (if available)
            "rule_category": "",
            "rule_subcategories": [],
            "vulnerability_classes": [],
            "cwe_names": [],
            "owasp_names": []
        })
        
        # Add rule information if available
        if finding.rule:
            finding_dict.update({
                "rule_category": finding.rule.category,
                "rule_subcategories": finding.rule.subcategories,
                "vulnerability_classes": finding.rule.vulnerability_classes,
                "cwe_names": finding.rule.cwe_names,
                "owasp_names": finding.rule.owasp_names
            })
            
        # Add assistant information if available
        if finding.assistant:
            if finding.assistant.guidance:
                finding_dict["guidance_summary"] = finding.assistant.guidance.summary
                finding_dict["guidance_instructions"] = finding.assistant.guidance.instructions
                
            if finding.assistant.autofix:
                finding_dict["autofix_code"] = finding.assistant.autofix.fix_code
                finding_dict["autofix_explanation"] = finding.assistant.autofix.explanation
                
            if finding.assistant.autotriage:
                finding_dict["autotriage_verdict"] = finding.assistant.autotriage.verdict
                finding_dict["autotriage_reason"] = finding.assistant.autotriage.reason
                
            if finding.assistant.component:
                finding_dict["component_tag"] = finding.assistant.component.tag
                finding_dict["component_risk"] = finding.assistant.component.risk
                
        return finding_dict

    def _add_severity_summary_table(self, pdf: FPDF, findings: List[Finding], font_family: str) -> None:
        """Add a table showing severity counts per repository."""
        # Group findings by repository and count severities
        repo_severity_counts = {}
        for finding in findings:
            repo = finding.get_repository_name() or "Unknown Repository"
            severity = finding.get_severity().upper() if finding.get_severity() else 'INFO'
            
            if repo not in repo_severity_counts:
                repo_severity_counts[repo] = {
                    'CRITICAL': 0,
                    'HIGH': 0,
                    'MEDIUM': 0,
                    'LOW': 0
                }
            
            if severity in repo_severity_counts[repo]:
                repo_severity_counts[repo][severity] += 1

        # Add table header
        pdf.ln(10)
        pdf.set_font(font_family, 'B', 12)
        pdf.cell(0, 10, "Open Findings by Repository and Severity", ln=True)
        pdf.ln(5)

        # Define column widths (total = 190)
        col_widths = {
            'repo': 70,
            'critical': 30,
            'high': 30,
            'medium': 30,
            'low': 30
        }

        # Add table headers
        pdf.set_font(font_family, 'B', 9)
        pdf.set_fill_color(240, 240, 240)  # Light gray background
        
        # Header row
        pdf.cell(col_widths['repo'], 8, "Repository", 1, 0, 'L', True)
        pdf.cell(col_widths['critical'], 8, "Critical", 1, 0, 'C', True)
        pdf.cell(col_widths['high'], 8, "High", 1, 0, 'C', True)
        pdf.cell(col_widths['medium'], 8, "Medium", 1, 0, 'C', True)
        pdf.cell(col_widths['low'], 8, "Low", 1, 1, 'C', True)

        # Add data rows
        pdf.set_font(font_family, '', 9)
        for repo, counts in repo_severity_counts.items():
            # Repository name might be long, so we'll handle wrapping
            repo_name = repo
            if len(repo_name) > 40:  # Truncate long names
                repo_name = repo_name[:37] + "..."
            
            pdf.cell(col_widths['repo'], 8, repo_name, 1, 0, 'L')
            pdf.cell(col_widths['critical'], 8, str(counts['CRITICAL']), 1, 0, 'C')
            pdf.cell(col_widths['high'], 8, str(counts['HIGH']), 1, 0, 'C')
            pdf.cell(col_widths['medium'], 8, str(counts['MEDIUM']), 1, 0, 'C')
            pdf.cell(col_widths['low'], 8, str(counts['LOW']), 1, 1, 'C')

        pdf.ln(10)

    def _generate_pdf_report(self, findings: List[Finding]) -> None:
        """Generate PDF report with charts and formatted findings."""
        logger.debug("Generating PDF report")
        try:
            pdf = FPDF()
            
            # Set up font handling for Unicode text
            font_family = 'DejaVu'
            use_unicode_font = False
            
            # First try with the DejaVu fonts we downloaded
            try:
                import os
                
                # Path to our embedded fonts
                fonts_dir = os.path.join(os.path.dirname(__file__), 'fonts')
                
                dejavu_regular = os.path.join(fonts_dir, 'DejaVuSans.ttf')
                dejavu_bold = os.path.join(fonts_dir, 'DejaVuSans-Bold.ttf')
                dejavu_italic = os.path.join(fonts_dir, 'DejaVuSans-Oblique.ttf')
                
                # Check if our font files exist
                if all(os.path.exists(f) for f in [dejavu_regular, dejavu_bold, dejavu_italic]):
                    pdf.add_font(font_family, '', dejavu_regular, uni=True)
                    pdf.add_font(font_family, 'B', dejavu_bold, uni=True)
                    pdf.add_font(font_family, 'I', dejavu_italic, uni=True)
                    use_unicode_font = True
                    logger.debug(f"Using embedded DejaVu Sans fonts: {dejavu_regular}")
                else:
                    logger.warning(f"Embedded DejaVu fonts not found in {fonts_dir}")
            except Exception as font_error:
                logger.warning(f"Error adding embedded DejaVu fonts: {font_error}")
            
            # Next try with a built-in Unicode font (recent versions of FPDF support this)
            if not use_unicode_font:
                try:
                    pdf.set_font('helvetica', '')
                    # If we get here, then modern FPDF version with built-in Unicode support
                    font_family = 'helvetica'
                    use_unicode_font = True
                    logger.debug("Using built-in Unicode font 'helvetica'")
                except Exception as e:
                    logger.debug(f"Built-in Unicode font not available: {e}")
                    font_family = 'Arial'
                    logger.info("Using standard fonts with text sanitization")
            
            # Add a handler to sanitize text, removing problematic Unicode characters
            def sanitize_text(text):
                if not text:
                    return ""
                # Replace smart quotes and other problematic characters
                replacements = {
                    '\u2018': "'", # Left single quote
                    '\u2019': "'", # Right single quote
                    '\u201c': '"', # Left double quote
                    '\u201d': '"', # Right double quote
                    '\u2013': '-', # En dash
                    '\u2014': '--', # Em dash
                    '\u00a0': ' ', # Non-breaking space
                    '\u2022': '*', # Bullet
                    '\u2026': '...', # Ellipsis
                    '\u2032': "'", # Prime (like apostrophe)
                    '\u2033': '"', # Double prime (like quotes)
                }
                for char, replacement in replacements.items():
                    text = text.replace(char, replacement)
                
                # Remove other non-Latin1 characters
                result = ""
                for char in text:
                    if ord(char) < 256:  # Latin-1 range
                        result += char
                    else:
                        result += '?' # Replace with question mark
                return result
            
            # Monkey patch the FPDF normalize_text method to use our sanitizing function
            original_normalize_text = pdf.normalize_text
            def safe_normalize_text(text):
                return original_normalize_text(sanitize_text(text))
            pdf.normalize_text = safe_normalize_text
            
            pdf.add_page()

            # Add title
            pdf.set_font(font_family, 'B', 16)
            pdf.cell(0, 10, sanitize_text(self.config.report_title), ln=True, align="C")

            # Add company logo if configured
            if self.config.company_logo and self.config.company_logo.exists():
                logger.debug(f"Adding company logo from {self.config.company_logo}")
                pdf.image(str(self.config.company_logo), x=10, y=10, w=30)

            # Add summary statistics
            pdf.set_font(font_family, 'B', 12)
            pdf.cell(0, 10, sanitize_text(f"Total Findings: {len(findings)}"), ln=True)
            generation_date = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')
            pdf.cell(0, 10, sanitize_text(f"Generated on: {generation_date}"), ln=True)

            # Add severity summary table
            if findings:
                self._add_severity_summary_table(pdf, findings, font_family)

            # Add severity distribution chart if enabled
            if self.config.include_charts and findings:
                logger.debug("Adding charts to PDF")
                self._add_charts_to_pdf(pdf, findings)

            if not findings:
                logger.debug("No findings to display in PDF")
                pdf.set_font(font_family, '', 10)
                pdf.cell(0, 10, sanitize_text("No security findings were identified."), ln=True)
            else:
                # Group findings by repository
                findings_by_repo = {}
                for finding in findings:
                    repo = finding.get_repository_name() or "Unknown Repository"
                    if repo not in findings_by_repo:
                        findings_by_repo[repo] = []
                    findings_by_repo[repo].append(finding)

                # Process each repository's findings
                for repo_name, repo_findings in findings_by_repo.items():
                    pdf.add_page()
                    
                    # Repository header
                    pdf.set_font(font_family, 'B', 14)
                    pdf.cell(0, 10, sanitize_text(f"Repository: {repo_name}"), ln=True)
                    
                    # Calculate severity counts for this repository
                    severity_counts = {
                        'CRITICAL': 0,
                        'HIGH': 0,
                        'MEDIUM': 0,
                        'LOW': 0,
                        'INFO': 0
                    }
                    
                    for finding in repo_findings:
                        severity = finding.get_severity().upper() if finding.get_severity() else 'INFO'
                        severity_counts[severity] = severity_counts.get(severity, 0) + 1
                    
                    # Display severity summary for this repository
                    pdf.set_font(font_family, 'B', 10)
                    pdf.cell(0, 8, sanitize_text("Severity Summary:"), ln=True)
                    pdf.set_font(font_family, '', 9)
                    
                    # Display counts with color-coded severity
                    for severity, count in severity_counts.items():
                        if count > 0:  # Only show severities that have findings
                            # Set color based on severity
                            if severity == 'CRITICAL':
                                pdf.set_text_color(139, 0, 0)  # Dark Red
                            elif severity == 'HIGH':
                                pdf.set_text_color(255, 0, 0)  # Red
                            elif severity == 'MEDIUM':
                                pdf.set_text_color(255, 140, 0)  # Orange
                            elif severity == 'LOW':
                                pdf.set_text_color(0, 128, 0)  # Green
                            else:
                                pdf.set_text_color(128, 128, 128)  # Gray
                            
                            pdf.cell(0, 5, sanitize_text(f"{severity}: {count} findings"), ln=True)
                    
                    # Reset text color to black
                    pdf.set_text_color(0, 0, 0)
                    
                    # Add a separator line
                    pdf.ln(5)
                    pdf.cell(0, 0, "", ln=True, border="T")
                    pdf.ln(5)
                    
                    # Display detailed findings for this repository
                    pdf.set_font(font_family, 'B', 12)
                    pdf.cell(0, 10, sanitize_text("Detailed Findings"), ln=True)
                    
                    pdf.set_font(font_family, '', 10)
                    for finding in repo_findings:
                        finding_dict = self._finding_to_dict(finding)
                        
                        # Basic finding information
                        pdf.set_font(font_family, 'B', 10)
                        pdf.cell(0, 10, sanitize_text(f"Finding: {finding_dict['check_id']}"), ln=True)
                        
                        pdf.set_font(font_family, '', 9)
                        pdf.multi_cell(0, 5, sanitize_text(
                            f"Severity: {finding_dict['severity'].upper() if finding_dict['severity'] else 'Unknown'}\n"
                            f"Status: {finding_dict['status'] or 'N/A'}\n"
                            f"State: {finding_dict['state'] or 'N/A'}\n"
                            f"Triage State: {finding_dict['triage_state'] or 'N/A'}\n"
                            f"File: {finding_dict['path']}:{finding_dict['line']}\n"
                        ), border=0)
                        
                        # Add SCA information if this is a dependency finding
                        if finding_dict['is_dependency']:
                            pdf.ln(2)
                            pdf.set_font(font_family, 'B', 9)
                            pdf.cell(0, 5, sanitize_text("Dependency Information:"), ln=True)
                            pdf.set_font(font_family, '', 9)
                            
                            sca_info = ""
                            if finding_dict['dependency_name']:
                                sca_info += f"Package: {finding_dict['dependency_name']}\n"
                            if finding_dict['dependency_version']:
                                sca_info += f"Version: {finding_dict['dependency_version']}\n"
                            if finding_dict['fixed_version']:
                                sca_info += f"Fixed in Version: {finding_dict['fixed_version']}\n"
                            if finding_dict['ecosystem']:
                                sca_info += f"Ecosystem: {finding_dict['ecosystem']}\n"
                            if finding_dict['cve_ids']:
                                sca_info += f"CVE IDs: {', '.join(finding_dict['cve_ids'])}\n"
                            if finding_dict['reachable'] is not None:
                                sca_info += f"Reachable: {'Yes' if finding_dict['reachable'] else 'No'}\n"
                            if finding_dict['reachability_details']:
                                sca_info += f"Reachability Details: {finding_dict['reachability_details']}\n"
                            if finding_dict['references']:
                                sca_info += f"References:\n"
                                for ref in finding_dict['references']:
                                    sca_info += f"  - {ref}\n"
                            
                            pdf.multi_cell(0, 5, sanitize_text(sca_info), border=0)
                        
                        # Add triage information if available
                        if any([finding_dict['triaged_at'], finding_dict['triage_comment'], finding_dict['triage_reason'], finding_dict['state_updated_at']]):
                            pdf.ln(2)
                            pdf.set_font(font_family, 'B', 9)
                            pdf.cell(0, 5, sanitize_text("Triage Information:"), ln=True)
                            pdf.set_font(font_family, '', 9)
                            
                            triage_info = ""
                            if finding_dict['triaged_at']:
                                triage_info += f"Triaged on: {finding_dict['triaged_at']}\n"
                            if finding_dict['triage_reason']:
                                triage_info += f"Triage Reason: {finding_dict['triage_reason']}\n"
                            if finding_dict['triage_comment']:
                                triage_info += f"Comment: {finding_dict['triage_comment']}\n"
                            if finding_dict['state_updated_at']:
                                triage_info += f"State Last Updated: {finding_dict['state_updated_at']}\n"
                            
                            pdf.multi_cell(0, 5, sanitize_text(triage_info), border=0)
                        
                        # Link to code if available
                        if finding_dict['line_of_code_url']:
                            pdf.set_text_color(0, 0, 255)  # Blue for links
                            pdf.set_font(font_family, 'U', 9)  # Underlined text for links
                            pdf.cell(0, 5, sanitize_text("View in repository"), link=finding_dict['line_of_code_url'])
                            pdf.ln()
                            
                            # Add Semgrep UI link if available
                            if finding_dict['semgrep_ui_url']:
                                pdf.cell(0, 5, sanitize_text("View in Semgrep UI"), link=finding_dict['semgrep_ui_url'])
                                pdf.ln()
                            
                            pdf.set_font(font_family, '', 9)  # Reset font
                            pdf.set_text_color(0, 0, 0)  # Reset to black
                        
                        # Message/Description
                        pdf.set_font(font_family, 'B', 9)
                        pdf.cell(0, 5, sanitize_text("Description:"), ln=True)
                        pdf.set_font(font_family, '', 9)
                        
                        # Make sure message is not None before using it
                        message = finding_dict.get('message', '')
                        if message:
                            pdf.multi_cell(0, 5, sanitize_text(message), border=0)
                        else:
                            pdf.multi_cell(0, 5, sanitize_text("No description available"), border=0)
                        
                        # Rule details if available
                        if finding.rule:
                            pdf.ln(2)
                            pdf.set_font(font_family, 'B', 9)
                            pdf.cell(0, 5, sanitize_text("Rule Details:"), ln=True)
                            pdf.set_font(font_family, '', 9)
                            
                            # Build rule details text
                            rule_details = ""
                            if finding.rule.category:
                                rule_details += f"Category: {finding.rule.category}\n"
                            
                            if finding.rule.subcategories:
                                rule_details += f"Subcategories: {', '.join(finding.rule.subcategories)}\n"
                            
                            if finding.rule.vulnerability_classes:
                                rule_details += f"Vulnerability Classes: {', '.join(finding.rule.vulnerability_classes)}\n"
                            
                            if finding.rule.cwe_names:
                                rule_details += f"CWE: {', '.join(finding.rule.cwe_names)}\n"
                            
                            if finding.rule.owasp_names:
                                rule_details += f"OWASP: {', '.join(finding.rule.owasp_names)}\n"
                            
                            if rule_details:
                                pdf.multi_cell(0, 5, sanitize_text(rule_details), border=0)
                            else:
                                pdf.multi_cell(0, 5, sanitize_text("No additional rule details available"), border=0)
                        
                        # Assistant guidance if available
                        if finding.assistant and hasattr(finding.assistant, 'guidance') and finding.assistant.guidance:
                            guidance_summary = getattr(finding.assistant.guidance, 'summary', None)
                            guidance_instructions = getattr(finding.assistant.guidance, 'instructions', None)
                            
                            if guidance_summary or guidance_instructions:
                                pdf.ln(2)
                                pdf.set_font(font_family, 'B', 9)
                                pdf.cell(0, 5, sanitize_text("Remediation Guidance:"), ln=True)
                                pdf.set_font(font_family, '', 9)
                                
                                if guidance_summary:
                                    pdf.multi_cell(0, 5, sanitize_text(f"Summary: {guidance_summary}"), border=0)
                                
                                if guidance_instructions:
                                    pdf.ln(1)
                                    pdf.set_font(font_family, 'I', 8)
                                    pdf.multi_cell(0, 4, sanitize_text(guidance_instructions), border=0)
                        
                        # Autofix if available  
                        if finding.assistant and hasattr(finding.assistant, 'autofix') and finding.assistant.autofix:
                            fix_code = getattr(finding.assistant.autofix, 'fix_code', None)
                            explanation = getattr(finding.assistant.autofix, 'explanation', None)
                            
                            if fix_code:
                                pdf.ln(2)
                                pdf.set_font(font_family, 'B', 9)
                                pdf.cell(0, 5, sanitize_text("Suggested Fix:"), ln=True)
                                
                                # Use the same font for code (no Courier in many Unicode fonts)
                                pdf.set_font(font_family, '', 8)  # Smaller size for code
                                pdf.multi_cell(0, 4, sanitize_text(fix_code), border=0)
                                
                                if explanation:
                                    pdf.set_font(font_family, 'I', 8)
                                    pdf.multi_cell(0, 4, sanitize_text(f"Note: {explanation}"), border=0)
                        
                        # Add separator between findings
                        pdf.ln(5)
                        pdf.cell(0, 0, "", ln=True, border="T")  # Horizontal line
                        pdf.ln(5)

            output_path = self.output_dir / "semgrep_report.pdf"
            pdf.output(str(output_path))
            logger.info(f"PDF report saved to {output_path}")
            
        except Exception as e:
            logger.error(f"Error in PDF generation: {e}")
            logger.exception("PDF generation error details")
            raise

    def _add_charts_to_pdf(self, pdf: FPDF, findings: List[Finding]) -> None:
        """Add visualization charts to the PDF report."""
        if not findings:
            logger.debug("Skipping chart generation - no findings")
            return  # Don't try to create charts if there are no findings
            
        try:
            # Convert findings to DataFrame for analysis
            data = [self._finding_to_dict(f) for f in findings]
            
            if not data:
                logger.debug("No data extracted for charts")
                return  # If we couldn't extract data, don't try to create charts
                
            df = pd.DataFrame(data)
            
            # Count severity levels for the chart
            severity_counts = df["severity"].value_counts().reset_index()
            severity_counts.columns = ["severity", "count"]
            
            if len(severity_counts) == 0:
                logger.debug("No severity data for charts")
                return  # If there's no severity data, don't try to create charts
            
            # Create severity distribution chart
            fig = px.pie(severity_counts, values="count", names="severity", 
                        title="Findings by Severity",
                        color_discrete_sequence=px.colors.qualitative.Set3)
            
            chart_path = self.output_dir / "severity_chart.png"
            fig.write_image(str(chart_path))
            logger.debug(f"Chart saved to {chart_path}")
            
            # Always start charts on a new page
            pdf.add_page()
            
            # Add chart to PDF - centered on the page
            page_width = pdf.w
            chart_width = 190  # Width of the chart in mm
            x_position = (page_width - chart_width) / 2  # Center horizontally
            pdf.image(str(chart_path), x=x_position, y=30, w=chart_width)  # Fixed y position from top
            
            # Clean up temporary chart file
            chart_path.unlink(missing_ok=True)
            
        except Exception as e:
            logger.error(f"Error generating charts: {e}")
            logger.exception("Chart generation error details")
            # Continue without charts rather than failing the whole report

    def _generate_csv_report(self, findings: List[Finding]) -> None:
        """Generate CSV report with all findings."""
        logger.debug("Generating CSV report")
        output_path = self.output_dir / "semgrep_findings.csv"
        
        try:
            # Define all the fields we want in our CSV
            fieldnames = [
                # Basic finding information
                "check_id", "path", "line", "message", "severity", "repository",
                
                # Links and IDs
                "finding_id", "line_of_code_url", "semgrep_ui_url",
                
                # Triage fields
                "state", "status", "triage_state", "triaged_at", "state_updated_at", "triage_comment", "triage_reason",
                
                # Optional fields
                "commit", "scan_date",
                
                # SCA-specific fields
                "is_dependency", "dependency_name", "dependency_version", "fixed_version", "ecosystem", "cve_ids", "references", "reachable", "reachability_details",
                
                # Rule fields
                "rule_category", "rule_subcategories", "vulnerability_classes", "cwe_names", "owasp_names",
                
                # Assistant guidance
                "guidance_summary", "guidance_instructions",
                
                # Assistant autofix 
                "autofix_code", "autofix_explanation",
                
                # Assistant autotriage
                "autotriage_verdict", "autotriage_reason",
                
                # Component
                "component_tag", "component_risk"
            ]
            
            with open(output_path, "w", newline="") as csvfile:
                writer = csv.DictWriter(
                    csvfile,
                    fieldnames=fieldnames
                )
                writer.writeheader()
                
                if not findings:
                    logger.debug("No findings to write to CSV")
                else:
                    for finding in findings:
                        writer.writerow(self._finding_to_dict(finding))
            
            logger.info(f"CSV report saved to {output_path}")
            
        except Exception as e:
            logger.error(f"Error in CSV generation: {e}")
            raise

    def _generate_excel_report(self, findings: List[Finding]) -> None:
        """Generate Excel report with findings and charts."""
        logger.debug("Generating Excel report")
        
        try:
            # Convert findings to a list of dictionaries for pandas
            data = [self._finding_to_dict(f) for f in findings]
            
            # Get the fieldnames list from the CSV report generation
            fieldnames = [
                # Basic finding information
                "check_id", "path", "line", "message", "severity", "repository",
                
                # Links and IDs
                "finding_id", "line_of_code_url", "semgrep_ui_url",
                
                # Triage fields
                "state", "status", "triage_state", "state_updated_at", "triage_comment", "triage_reason",
                
                # Optional fields
                "commit", "scan_date",
                
                # SCA-specific fields
                "is_dependency", "dependency_name", "dependency_version", "fixed_version", "ecosystem", "cve_ids", "references", "reachable", "reachability_details",
                
                # Rule fields
                "rule_category", "rule_subcategories", "vulnerability_classes", "cwe_names", "owasp_names",
                
                # Assistant guidance
                "guidance_summary", "guidance_instructions",
                
                # Assistant autofix 
                "autofix_code", "autofix_explanation",
                
                # Assistant autotriage
                "autotriage_verdict", "autotriage_reason",
                
                # Component
                "component_tag", "component_risk"
            ]
            
            df = pd.DataFrame(data)
            output_path = self.output_dir / "semgrep_findings.xlsx"
            
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                # Create severity summary by repository
                if not df.empty:
                    summary_data = []
                    for repo in df['repository'].unique():
                        repo_findings = df[df['repository'] == repo]
                        severity_counts = {
                            'Repository': repo,
                            'Critical': len(repo_findings[repo_findings['severity'].str.upper() == 'CRITICAL']),
                            'High': len(repo_findings[repo_findings['severity'].str.upper() == 'HIGH']),
                            'Medium': len(repo_findings[repo_findings['severity'].str.upper() == 'MEDIUM']),
                            'Low': len(repo_findings[repo_findings['severity'].str.upper() == 'LOW'])
                        }
                        summary_data.append(severity_counts)
                    
                    summary_df = pd.DataFrame(summary_data)
                else:
                    summary_df = pd.DataFrame(columns=['Repository', 'Critical', 'High', 'Medium', 'Low'])

                # Write the summary sheet first
                summary_df.to_excel(writer, sheet_name="Open Findings by Repository", index=False)
                
                # Format the summary sheet
                workbook = writer.book
                summary_sheet = writer.sheets["Open Findings by Repository"]
                
                # Apply formatting to the summary sheet
                for column in range(len(summary_df.columns)):
                    max_length = 0
                    column_letter = get_column_letter(column + 1)
                    
                    # Find the maximum length in the column
                    for row in range(len(summary_df) + 1):  # +1 for header
                        cell = summary_sheet[f"{column_letter}{row + 1}"]
                        max_length = max(max_length, len(str(cell.value)))
                    
                    # Set the column width
                    adjusted_width = (max_length + 2)
                    summary_sheet.column_dimensions[column_letter].width = adjusted_width

                # Write the detailed findings
                if df.empty:
                    pd.DataFrame(columns=fieldnames).to_excel(writer, sheet_name="Findings", index=False)
                else:
                    columns_to_include = [col for col in fieldnames if col in df.columns]
                    df = df[columns_to_include]
                    df.to_excel(writer, sheet_name="Findings", index=False)
                    
                    if self.config.include_charts:
                        pivot = pd.pivot_table(
                            df,
                            values="severity",
                            index="severity",
                            aggfunc="count"
                        )
                        pivot.to_excel(writer, sheet_name="Charts")
            
            logger.info(f"Excel report saved to {output_path}")
            
        except Exception as e:
            logger.error(f"Error in Excel generation: {e}")
            raise 